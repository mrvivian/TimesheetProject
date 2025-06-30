"""
Timesheet Template Generator for Power BI
=========================================

This script generates a user-friendly Excel timesheet template optimized for Power BI reporting.
Features include:
- Separate worksheets for each month (June 2025 - May 2026)
- Multiple rows per day for different projects
- Pre-populated weekdays with Part Time Work schedule
- Weekend detection and automatic population
- Time formatting for easy data entry
- Automatic calculation of total work hours
- Clean structure for Power BI import

Author: Vivian Ferguson
Date: June 2025
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle
from datetime import datetime, timedelta
import calendar
import os

def create_timesheet_template():
    """Generate the timesheet template with all features"""
    
    # Define columns for the main sheet
    timesheet_columns = [
        'Date',
        'Employee Name',
        'Project',
        'Task Description',
        'Work Start',
        'Work End',
        'Break Start',
        'Break End',
        'Gym Start',
        'Gym End',
        'Commute Start',
        'Commute End',
        'Total Work Hours'
    ]

    # Example projects
    projects = ['Part Time Work', 'Self Employment', 'Vivian Portfolio', 'Break', 'Leave', 'Weekend']

    # Create a new workbook
    wb = Workbook()

    # Remove the default sheet
    wb.remove(wb.active)

    # Create Projects sheet first
    ws_projects = wb.create_sheet('Projects')

    # Add projects to Projects sheet
    for proj in projects:
        ws_projects.append([proj])

    # Create sheets for each month (June 2025 to May 2026)
    months = [
        ('June 2025', 2025, 6),
        ('July 2025', 2025, 7),
        ('August 2025', 2025, 8),
        ('September 2025', 2025, 9),
        ('October 2025', 2025, 10),
        ('November 2025', 2025, 11),
        ('December 2025', 2025, 12),
        ('January 2026', 2026, 1),
        ('February 2026', 2026, 2),
        ('March 2026', 2026, 3),
        ('April 2026', 2026, 4),
        ('May 2026', 2026, 5)
    ]

    for month_name, year, month in months:
        # Create worksheet for this month
        ws = wb.create_sheet(month_name)
        
        # Add headers
        ws.append(timesheet_columns)
        
        # Get the number of days in this month
        num_days = calendar.monthrange(year, month)[1]
        
        # Pre-populate dates for this month with multiple rows per day
        row_num = 2
        for day in range(1, num_days + 1):
            current_date = datetime(year, month, day)
            
            # Check if it's a weekend
            is_weekend = current_date.weekday() >= 5  # Saturday = 5, Sunday = 6
            
            if is_weekend:
                # Weekend - just one row with Weekend
                ws.append([
                    current_date.strftime('%Y-%m-%d'),
                    'Vivian',
                    'Weekend',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    None
                ])
                row_num += 1
            else:
                # Weekday - create multiple rows for different activities
                
                # Row 1: Part Time Work (pre-populated)
                ws.append([
                    current_date.strftime('%Y-%m-%d'),
                    'Vivian',
                    'Part Time Work',
                    '',
                    '08:00',
                    '13:00',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    None
                ])
                row_num += 1
                
                # Row 2: Self Employment (blank for user to fill)
                ws.append([
                    current_date.strftime('%Y-%m-%d'),
                    'Vivian',
                    'Self Employment',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    None
                ])
                row_num += 1
                
                # Row 3: Additional entry (blank for user to fill)
                ws.append([
                    current_date.strftime('%Y-%m-%d'),
                    'Vivian',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    None
                ])
                row_num += 1
        
        # Set column widths
        for i, col in enumerate(timesheet_columns, 1):
            ws.column_dimensions[get_column_letter(i)].width = 18
        
        # Set time format for time columns
        time_columns = ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']
        time_style = NamedStyle(name="time_style", number_format='HH:MM')
        for col in time_columns:
            for row in range(2, row_num + 1):
                ws[f"{col}{row}"].style = time_style
        
        # Add formula for Total Work Hours (M) - simplified to avoid errors
        for row in range(2, row_num + 1):
            # Simple formula that calculates work hours minus break time
            ws[f'M{row}'] = f'=IF(AND(E{row}<>"",F{row}<>""),F{row}-E{row}-IF(AND(G{row}<>"",H{row}<>""),H{row}-G{row},0),"")'

    return wb

def main():
    """Main function to generate the timesheet template"""
    print("Generating Timesheet Template for Power BI...")
    
    try:
        # Create the template
        wb = create_timesheet_template()
        
        # Save the workbook
        output_file = 'Timesheet_Template_For_PowerBI.xlsx'
        wb.save(output_file)
        
        print(f"✅ Successfully generated: {output_file}")
        print("\n📋 Template Features:")
        print("• 12 monthly worksheets (June 2025 - May 2026)")
        print("• Multiple rows per day for different projects")
        print("• Pre-populated weekdays with Part Time Work (08:00-13:00)")
        print("• Weekend detection and automatic population")
        print("• Time formatting for easy data entry")
        print("• Automatic calculation of total work hours")
        print("• Clean structure optimized for Power BI import")
        print("\n📝 Usage Instructions:")
        print("1. Open the Excel file")
        print("2. Navigate to any month tab")
        print("3. Enter your time data in the appropriate columns")
        print("4. Import into Power BI for reporting")
        print("\n🔧 Optional: Add dropdowns manually in Excel:")
        print("• Select Project column (C)")
        print("• Go to Data > Data Validation")
        print("• Choose 'List' and enter: Part Time Work,Self Employment,Vivian Portfolio,Break,Leave,Weekend")
        
    except Exception as e:
        print(f"❌ Error generating template: {e}")

if __name__ == "__main__":
    main() 